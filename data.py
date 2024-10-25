import tkinter as tk
from tkinter import ttk
from tkcalendar import DateEntry
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import os

# Create or load the Excel workbook
file_name = 'names.xlsx'
if os.path.exists(file_name):
    workbook = load_workbook(file_name)
    sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = 'ID'
    sheet['B1'] = 'Name'
    sheet['C1'] = 'Date of Birth'
    sheet['D1'] = 'Date'
    sheet['E1'] = 'Address'
    sheet['F1'] = 'Mobile Number'
    sheet['G1'] = 'Age'
    sheet['H1'] = 'Medical Problem'

# Function to save the details to the Excel sheet
def save_details():
    name = entry_name.get()
    dob = entry_dob.get_date()
    date = entry_date.get_date()
    address = entry_address.get()
    mobile = entry_mobile.get()
    age = entry_age.get()
    medical_problem = entry_medical_problem.get()
    
    if name and dob and date and address and mobile and age and medical_problem:
        # Determine the next ID (row number)
        next_id = sheet.max_row + 1
        sheet.append([next_id, name, dob.strftime('%Y-%m-%d'), date.strftime('%Y-%m-%d'), address, mobile, age, medical_problem])
        workbook.save(file_name)
        status_label.config(text=f"The details for '{name}' have been added.")
    else:
        status_label.config(text="Please fill in all fields.")

# Function to clear the input fields for the next entry
def next_entry():
    entry_name.delete(0, tk.END)
    entry_dob.set_date('')
    entry_date.set_date('')
    entry_address.delete(0, tk.END)
    entry_mobile.delete(0, tk.END)
    entry_age.delete(0, tk.END)
    entry_medical_problem.delete(0, tk.END)
    status_label.config(text="Ready for next entry.")

# Function to search for a name and display all entries for that person
def search_name():
    name_to_search = entry_search_name.get()
    if not name_to_search:
        status_label.config(text="Please enter a name to search.")
        return
    
    # Clear previous search results
    for item in tree.get_children():
        tree.delete(item)
    
    found = False
    last_visit_date = None
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from the second row
        if row[1] == name_to_search:
            tree.insert('', tk.END, values=row)
            found = True
            last_visit_date = row[3]  # Date of visit
    
    if not found:
        status_label.config(text=f"No details found for '{name_to_search}'.")
        last_visit_label.config(text="")
    else:
        last_visit_label.config(text=f"Last Visit Date: {last_visit_date}")

# Function to convert the data to a PDF
def convert_to_pdf():
    name_to_search = entry_search_name.get()
    pdf_name = entry_pdf_name.get()
    
    if not name_to_search:
        status_label.config(text="Please enter a name to search before generating PDF.")
        return

    if not pdf_name:
        pdf_name = f'{name_to_search}_details'
    
    pdf_file = f'{pdf_name}.pdf'
    c = canvas.Canvas(pdf_file, pagesize=letter)
    width, height = letter

    y = height - 50  # Start position from top

    # Header
    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, y, f"Details for {name_to_search}")
    y -= 30

    # Fetch data and add to PDF
    found = False
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == name_to_search:
            c.setFont("Helvetica", 12)
            c.drawString(50, y, f"ID: {row[0]}")
            y -= 20
            c.drawString(50, y, f"Name: {row[1]}")
            y -= 20
            c.drawString(50, y, f"Date of Birth: {row[2]}")
            y -= 20
            c.drawString(50, y, f"Date: {row[3]}")
            y -= 20
            c.drawString(50, y, f"Address: {row[4]}")
            y -= 20
            c.drawString(50, y, f"Mobile Number: {row[5]}")
            y -= 20
            c.drawString(50, y, f"Age: {row[6]}")
            y -= 20
            c.drawString(50, y, f"Medical Problem: {row[7]}")
            y -= 30
            found = True
            break
    
    if not found:
        c.setFont("Helvetica", 12)
        c.drawString(50, y, f"No details found for '{name_to_search}'.")

    c.save()
    status_label.config(text=f"PDF generated as '{pdf_file}'.")

# Create the GUI window
root = tk.Tk()
root.title("Details Entry")

# Set the size of the window (width x height)
root.geometry('800x600')

# Title label at the top
tk.Label(root, text="SAVANT'S", font=('Arial', 24, 'bold')).grid(row=0, column=0, columnspan=4, pady=10)
tk.Label(root, text="WELLNESS CENTRE", font=('Arial', 14, 'bold')).grid(row=1, column=0, columnspan=4, pady=10)

# Create and place the widgets using grid layout
tk.Label(root, text="Enter Name:").grid(row=2, column=0, padx=10, pady=5, sticky='e')
entry_name = tk.Entry(root, font=('Arial', 12))
entry_name.grid(row=2, column=1, padx=10, pady=5, sticky='w')

tk.Label(root, text="Date of Birth:").grid(row=3, column=0, padx=10, pady=5, sticky='e')
entry_dob = DateEntry(root, font=('Arial', 12), date_pattern='yyyy-mm-dd')
entry_dob.grid(row=3, column=1, padx=10, pady=5, sticky='w')

tk.Label(root, text="Date:").grid(row=4, column=0, padx=10, pady=5, sticky='e')
entry_date = DateEntry(root, font=('Arial', 12), date_pattern='yyyy-mm-dd')
entry_date.grid(row=4, column=1, padx=10, pady=5, sticky='w')

tk.Label(root, text="Address:").grid(row=5, column=0, padx=10, pady=5, sticky='e')
entry_address = tk.Entry(root, font=('Arial', 12))
entry_address.grid(row=5, column=1, padx=10, pady=5, sticky='w')

tk.Label(root, text="Mobile Number:").grid(row=6, column=0, padx=10, pady=5, sticky='e')
entry_mobile = tk.Entry(root, font=('Arial', 12))
entry_mobile.grid(row=6, column=1, padx=10, pady=5, sticky='w')

tk.Label(root, text="Age:").grid(row=7, column=0, padx=10, pady=5, sticky='e')
entry_age = tk.Entry(root, font=('Arial', 12))
entry_age.grid(row=7, column=1, padx=10, pady=5, sticky='w')

tk.Label(root, text="Medical Problem:").grid(row=8, column=0, padx=10, pady=5, sticky='e')
entry_medical_problem = tk.Entry(root, font=('Arial', 12))
entry_medical_problem.grid(row=8, column=1, padx=10, pady=5, sticky='w')

save_button = tk.Button(root, text="Save", command=save_details, font=('Arial', 12))
save_button.grid(row=9, column=0, columnspan=2, pady=10)

next_button = tk.Button(root, text="Next", command=next_entry, font=('Arial', 12))
next_button.grid(row=10, column=0, columnspan=2, pady=10)

# Search Section
tk.Label(root, text="Search Name:").grid(row=2, column=2, padx=10, pady=5, sticky='e')
entry_search_name = tk.Entry(root, font=('Arial', 12))
entry_search_name.grid(row=2, column=3, padx=10, pady=5, sticky='w')

search_button = tk.Button(root, text="Search", command=search_name, font=('Arial', 12))
search_button.grid(row=3, column=2, columnspan=2, pady=10)

# PDF Conversion Section
tk.Label(root, text="PDF Name:").grid(row=4, column=2, padx=10, pady=5, sticky='e')
entry_pdf_name = tk.Entry(root, font=('Arial', 12))
entry_pdf_name.grid(row=4, column=4, padx=10, pady=5, sticky='w')

convert_button = tk.Button(root, text="Convert to PDF", command=convert_to_pdf, font=('Arial', 12))
convert_button.grid(row=5, column=2, columnspan=2, pady=10)

# Results display area on the right side
columns = ['ID', 'Name', 'Date of Birth', 'Date', 'Address', 'Mobile Number', 'Age', 'Medical Problem']
tree = ttk.Treeview(root, columns=columns, show='headings', height=15)
tree.grid(row=6, column=2, columnspan=2, padx=10, pady=5, sticky='nsew')

# Define column headings
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=100, anchor='w')

# Last Visit Date Label
last_visit_label = tk.Label(root, text="", font=('Arial', 12))
last_visit_label.grid(row=10, column=2, columnspan=2, pady=10)

status_label = tk.Label(root, text="", font=('Arial', 10))
status_label.grid(row=11, column=2, columnspan=2, pady=20)

# Configure grid weights for resizing
root.grid_rowconfigure(6, weight=1)
root.grid_columnconfigure(2, weight=1)

# Start the GUI event loop
root.mainloop()
